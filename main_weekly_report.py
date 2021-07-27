from openpyxl import Workbook
from openpyxl import load_workbook
from weekly_report_files.py import ZhouBaoSource
#from print_date_of_week import DateOfWeek

zb_source_files = ZhouBaoSource()
zb_week_date1 = zb_source_files.work_week.get_week_date()
zb_week_date2 = zb_source_files.work_week.get_week_date_one_day_ago()
zb_week_date3 = zb_source_files.work_week.get_week_date_two_day_ago()
zb_week_date4 = zb_source_files.work_week.get_week_date_three_day_ago()
current_month_dir = zb_source_files.current_month_dir
zb_source_files_sheets = zb_source_files.work_week.sheet_name_date()
zb_bian_jie_zhi = zb_source_files.work_week.bian_jie_zhi

hui_zong = []
hui_zong_ph = []
hui_zong_ydjy = []
hui_zong_ydph = []
hui_zong_fk = []
hui_zong_qs = []
hui_zong_black = []

#excel写入数据
def write_table_into_sheet(sheet_name, zb_tables, start_position = [1, 1]):
    current_row = start_position[0]
    current_column = start_position[1]

    for i in zb_tables:
        for j in i:
            sheet_name.cell(row = current_row, column = current_column).value = j
            current_column += 1
        current_row += 1
        current_column = start_position[1]

#周报求和
def zhou_bao_qiu_he(hui_zong):
    wb_hui_zong = load_workbook("tong_ji_biao_ge.xlsx")

    for i, j in zip(wb_hui_zong.sheetnames, hui_zong):
        ws_hui_zong = wb_hui_zong[i]
        write_table_into_sheet(ws_hui_zong, j)

    wb_hui_zong.save("tong_ji_biao_ge.xlsx")

#账务平衡
def zhang_wu_ping_heng(wb_ph_name, ws_ph_names, zb_w1, zb_w3, hui_zong_ph):
    wb_ph = load_workbook(wb_ph_name)
    cha_yi_column = 6

    for i, q, o in zip(ws_ph_names, zb_w1, zb_w3):
        ws_ph = wb_ph[i]
        j = ws_ph.max_row
        hui_zong_ph_sheet = [0] * 4
        hui_zong_ph_sheet[0] = q
        hui_zong_ph_sheet[1] = o
        hui_zong_ph_sheet[2] += (j - 2)

        while j >= 1 and ws_ph.cell(row = j, column = cha_yi_column).value != 0:
            hui_zong_ph_sheet[3] += 1
            j -= 1

        hui_zong_ph.append(hui_zong_ph_sheet)

#异地交易
def yi_di_jiao_yi(wb_ydjy_name, ws_ydjy_names, zb_w1, zb_w2, hui_zong_ydjy):
    wb_ydjy = load_workbook(wb_ydjy_name)

    for i, j, q in zip(ws_ydjy_names, zb_w1, zb_w2):
        ws_ydjy = wb_ydjy[i]
        ying_shou_row = 25
        ying_fu_row = ws_ydjy.max_row
        bi_shu_column = 4
        jin_e_column = 5
        hui_zong_ydjy_sheet = [0] * 6

        hui_zong_ydjy_sheet[0] = j
        hui_zong_ydjy_sheet[1] = q
        hui_zong_ydjy_sheet[2] = int(ws_ydjy.cell(row = ying_shou_row, column = bi_shu_column).value[7:12].replace(",", ""))
        hui_zong_ydjy_sheet[3] = float(ws_ydjy.cell(row = ying_shou_row, column = jin_e_column).value[7:].replace(",", ""))
        hui_zong_ydjy_sheet[4] = int(ws_ydjy.cell(row = ying_fu_row, column = bi_shu_column).value[7:13].replace(",", ""))
        hui_zong_ydjy_sheet[5] = float(ws_ydjy.cell(row = ying_fu_row, column = jin_e_column).value[7:].replace(",", ""))

        hui_zong_ydjy.append(hui_zong_ydjy_sheet)

#异地平衡
def yi_di_ping_heng(wb_ydph_name, ws_ydph_names, zb_w1, zb_w4, hui_zong_ydph):
    wb_ydph = load_workbook(wb_ydph_name)

    for i, j, p in zip(ws_ydph_names, zb_w1, zb_w4):
        ws_ydph = wb_ydph[i]
        jin_e_row = 3
        jin_e_column = 2
        hui_zong_ydph_sheet = [0] * 6
        hui_zong_ydph_sheet[0] = j
        hui_zong_ydph_sheet[1] = p

        for q in range(0, 4):
            tmp_q = q + 2
            hui_zong_ydph_sheet[tmp_q] = ws_ydph.cell(row = jin_e_row, column = jin_e_column + q).value

        hui_zong_ydph.append(hui_zong_ydph_sheet)

#加油站文件接收入库情况
def wen_jian_ru_ku(wb_wjrk_name, ws_wjrk_names, zb_w1, hui_zong_black):
    wb_wjrk = load_workbook(wb_wjrk_name)

    for i, j in zip(ws_wjrk_names, zb_w1):
        ws_wjrk = wb_wjrk[i]
        wei_xia_zai = 0
        hui_zong_wjrk_sheet = [0] * 3
        hui_zong_wjrk_sheet[0] = j
        for q in range(3, ws_wjrk.max_row + 1):
            if ws_wjrk.row_dimensions[q].hidden == False:
                wei_xia_zai += 1

        hui_zong_wjrk_sheet[1] = wei_xia_zai // 2
        hui_zong_wjrk_sheet[2] = wei_xia_zai // 2

        hui_zong_black.append(hui_zong_wjrk_sheet)

#发卡网点异常流水
def fa_ka(wb_fk_name, ws_fk_names, zb_w1, hui_zong_fk):
    wb_fk = load_workbook(wb_fk_name)

    for i, q in zip(ws_fk_names, zb_w1):
        l_tag = True
        jin_e_column = 6
        chu_li_column = 14
        ws_fk = wb_fk[i]
        hui_zong_fk_sheet = [0] * 9
        hui_zong_fk_sheet[0] = q
        hui_zong_fk_sheet[1] = 0
        hui_zong_fk_sheet[5] = 0
        j = 8
        while j <= ws_fk.max_row:
            #已日结
            if ws_fk.cell(row = j, column = jin_e_column).value != None and l_tag:
                if ws_fk.cell(row = j, column = chu_li_column).value[0] == "已":
                    hui_zong_fk_sheet[6] += 1
                    hui_zong_fk_sheet[8] += float(ws_fk.cell(row = j, column = jin_e_column).value)
                hui_zong_fk_sheet[2] += 1
                hui_zong_fk_sheet[4] += float(ws_fk.cell(row = j, column = jin_e_column).value)
            elif ws_fk.cell(row = j, column = jin_e_column).value == None and l_tag:
                j += 3
                l_tag = False
            #X流水
            elif not l_tag:
                if ws_fk.cell(row = j, column = jin_e_column).value != None:
                    if ws_fk.cell(row = j, column = chu_li_column).value[0] == "已":
                        hui_zong_fk_sheet[7] += 1
                        hui_zong_fk_sheet[8] += float(ws_fk.cell(row = j, column = jin_e_column).value)
                    hui_zong_fk_sheet[3] += 1
                    hui_zong_fk_sheet[4] += float(ws_fk.cell(row = j, column = jin_e_column).value)
            j += 1

        hui_zong_fk.append(hui_zong_fk_sheet)

#清算异常流水
def qing_suan(wb_qs_name, ws_qs_names, zb_w1, hui_zong_qs):
    wb_qs = load_workbook(wb_qs_name)

    for i, q in zip(ws_qs_names, zb_w1):
        l_tag = True
        jin_e_column = 10
        chu_li_column = 36
        ws_qs = wb_qs[i]
        hui_zong_qs_sheet = [0] * 9
        hui_zong_qs_sheet[0] = q
        j = 4
        while j <= ws_qs.max_row:
            if ws_qs.cell(row = j, column = jin_e_column).value != None and l_tag:
                if ws_qs.cell(row = j, column = chu_li_column).value == "已处理":
                    hui_zong_qs_sheet[5] += 1
                    hui_zong_qs_sheet[7] += float(ws_qs.cell(row = j, column = jin_e_column).value)
                hui_zong_qs_sheet[1] += 1
                hui_zong_qs_sheet[3] += float(ws_qs.cell(row = j, column = jin_e_column).value)
            elif ws_qs.cell(row = j, column = jin_e_column).value == None and l_tag:
                j += 3
                jin_e_column += 1
                chu_li_column += 3
                l_tag = False
            elif not l_tag:
                if ws_qs.cell(row = j, column = jin_e_column).value != None:
                    if ws_qs.cell(row = j, column = chu_li_column).value == "已处理":
                        hui_zong_qs_sheet[6] += 1
                        hui_zong_qs_sheet[8] += float(ws_qs.cell(row = j, column = jin_e_column).value)
                    hui_zong_qs_sheet[2] += 1
                    hui_zong_qs_sheet[4] += float(ws_qs.cell(row = j, column = jin_e_column).value)
            j += 1

        hui_zong_qs.append(hui_zong_qs_sheet)

if zb_source_files.work_week.diferent_month:
    qing_suan(zb_source_files.last_week_qs, zb_source_files_sheets[:zb_bian_jie_zhi], zb_week_date1[:zb_bian_jie_zhi], hui_zong_qs)
    qing_suan(zb_source_files.current_week_qs, zb_source_files_sheets[zb_bian_jie_zhi:], zb_week_date1[zb_bian_jie_zhi:], hui_zong_qs)
    fa_ka(zb_source_files.last_week_fk, zb_source_files_sheets[:zb_bian_jie_zhi], zb_week_date1[:zb_bian_jie_zhi], hui_zong_fk)
    fa_ka(zb_source_files.current_week_fk, zb_source_files_sheets[zb_bian_jie_zhi:], zb_week_date1[zb_bian_jie_zhi:], hui_zong_fk)
    zhang_wu_ping_heng(zb_source_files.last_week_ph, zb_source_files_sheets[:zb_bian_jie_zhi], zb_week_date1[:zb_bian_jie_zhi], zb_week_date3[:zb_bian_jie_zhi], hui_zong_ph)
    zhang_wu_ping_heng(zb_source_files.current_week_ph, zb_source_files_sheets[zb_bian_jie_zhi:], zb_week_date1[zb_bian_jie_zhi:], zb_week_date3[zb_bian_jie_zhi:], hui_zong_ph)
    wen_jian_ru_ku(zb_source_files.last_week_black, zb_source_files_sheets[:zb_bian_jie_zhi], zb_week_date1[:zb_bian_jie_zhi], hui_zong_black)
    wen_jian_ru_ku(zb_source_files.current_week_black, zb_source_files_sheets[zb_bian_jie_zhi:], zb_week_date1[zb_bian_jie_zhi:], hui_zong_black)
    yi_di_jiao_yi(zb_source_files.last_week_ydjy, zb_source_files_sheets[:zb_bian_jie_zhi], zb_week_date1[:zb_bian_jie_zhi], zb_week_date2[:zb_bian_jie_zhi], hui_zong_ydjy)
    yi_di_jiao_yi(zb_source_files.current_week_ydjy, zb_source_files_sheets[zb_bian_jie_zhi:], zb_week_date1[zb_bian_jie_zhi:], zb_week_date2[zb_bian_jie_zhi:], hui_zong_ydjy)
    yi_di_ping_heng(zb_source_files.last_week_ydph, zb_source_files_sheets[:zb_bian_jie_zhi], zb_week_date1[:zb_bian_jie_zhi], zb_week_date4[:zb_bian_jie_zhi], hui_zong_ydph)
    yi_di_ping_heng(zb_source_files.current_week_ydph, zb_source_files_sheets[zb_bian_jie_zhi:], zb_week_date1[zb_bian_jie_zhi:], zb_week_date4[zb_bian_jie_zhi:], hui_zong_ydph)
else:
    qing_suan(zb_source_files.current_week_qs, zb_source_files_sheets[:zb_bian_jie_zhi], zb_week_date1[:zb_bian_jie_zhi], hui_zong_qs)
    fa_ka(zb_source_files.current_week_fk, zb_source_files_sheets[:zb_bian_jie_zhi], zb_week_date1[:zb_bian_jie_zhi], hui_zong_fk)
    zhang_wu_ping_heng(zb_source_files.current_week_ph, zb_source_files_sheets[:zb_bian_jie_zhi], zb_week_date1[:zb_bian_jie_zhi], zb_week_date3[:zb_bian_jie_zhi], hui_zong_ph)
    wen_jian_ru_ku(zb_source_files.current_week_black, zb_source_files_sheets[:zb_bian_jie_zhi], zb_week_date1[:zb_bian_jie_zhi], hui_zong_black)
    yi_di_jiao_yi(zb_source_files.current_week_ydjy, zb_source_files_sheets[:zb_bian_jie_zhi], zb_week_date1[:zb_bian_jie_zhi], zb_week_date2[:zb_bian_jie_zhi], hui_zong_ydjy)
    yi_di_ping_heng(zb_source_files.current_week_ydph, zb_source_files_sheets[:zb_bian_jie_zhi], zb_week_date1[:zb_bian_jie_zhi], zb_week_date4[:zb_bian_jie_zhi], hui_zong_ydph)

hui_zong.append(hui_zong_qs)
hui_zong.append(hui_zong_fk)
hui_zong.append(hui_zong_ph)
hui_zong.append(hui_zong_black)
hui_zong.append(hui_zong_ydjy)
hui_zong.append(hui_zong_ydph)

zhou_bao_qiu_he(hui_zong)
