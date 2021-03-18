from openpyxl import Workbook
from openpyxl import load_workbook
from print_date_of_week import DateOfWeek
from zhou_bao_files import ZhouBaoSource

#wb_week = Workbook()
#ws = wb_week.active
wb_week = load_workbook("统计表格.xlsx")
ws_week = wb_week.active
ws_week.max_row
ws_week.max_column

def create_new_table(self):
'''

ws['A1'] = 42
ws.append([1, 2, 3])
d = ws.cell(row=4, column=2, value=10)
d = ws.cell(row=3, column=2).value = "这是第三行, 第二列"
ws.append([1, 2, 3])
cell_range = ws['A1' : 'C2']
col_range = ws['C' : 'D']
for row in cell_range:
	for cell in row:
		print(cell.value)

for row in ws.values:
	for cell in row:
		print(cell)

for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
	print(row)

#tuple(ws.rows)
#tuple(ws.columns)

wb_week.save("sample.xlsx")
'''

def zhang_wu_ping_heng():
	hui_zong_zwph = []
	current_hui_zong = [0] * 2

	for i in 表格: 
		j = ws_zwph.max_row
		current_hui_zong[0] += (j - 2)

		while j >= 1 and j列.相差金額 != 0:
			current_hui_zong[1] += 1
			j -= 1

		hui_zong_zwph.append(current_hui_zong)

def yi_di_qing_suan():
	hui_zong_ydqs = []
	current_hui_zong = [0] * 4

	for i in 表格:
		current_hui_zong[0] = values
		current_hui_zong[1] = values
		current_hui_zong[2] = values
		current_hui_zong[3] = values
		hui_zong_ydqs.append(current_hui_zong)

def yi_di_ping_heng():
	hui_zong_ydph = []
	current_hui_zong = [0] * 4

	for i in 表格:
		for j in range(0, 4):
			current_hui_zong[j] = values

		hui_zong_ydph.append(current_hui_zong)
