from openpyxl import Workbook
from openpyxl import load_workbook
from zhou_bao_files import ZhouBaoSource
#from print_date_of_week import DateOfWeek

zb_source_files = ZhouBaoSource()
current_month_dir = zb_source_files.current_month_dir
zb_source_files_sheets = zb_source_files.work_week.sheet_name_date()
zb_bian_jie_zhi = zb_source_files.work_week.bian_jie_zhi

hui_zong_ph = []
hui_zong_ydjy = []
hui_zong_ydph = []
hui_zong_fk = []
hui_zong_qs = []
hui_zong_balck = []

#账务平衡
def zhang_wu_ping_heng(wb_ph_name, ws_ph_names, hui_zong_ph):
    wb_ph = load_workbook(wb_ph_name)
    cha_yi_column = 6

    for i in ws_ph_names:
        ws_ph = wb_ph[i]
        j = ws_ph.max_row
        hui_zong_ph_sheet = [0] * 2
        hui_zong_ph_sheet[0] += (j - 2)

        while j >= 1 and ws_ph.cell(row = j, column = cha_yi_column).value != 0:
            hui_zong_ph_sheet[1] += 1
            j -= 1

        hui_zong_ph.append(hui_zong_ph_sheet)

#异地交易
def yi_di_jiao_yi(wb_ydjy_name, ws_ydjy_names, hui_zong_ydjy):
    wb_ydjy = load_workbook(wb_ydjy_name)

    for i, j, q in zip(ws_ydjy_names, zb_source_files.work_week.get_week_date(), zb_source_files.work_week.get_week_date_one_day_ago()):
        ws_ydjy = wb_ydjy[i]
        ying_shou_row = 25
        ying_fu_row = ws_ydjy.max_row
        bi_shu_column = 4
        jin_e_column = 5
        hui_zong_ydjy_sheet = [0] * 6

        hui_zong_ydjy_sheet[0] = j
        hui_zong_ydjy_sheet[1] = q
        hui_zong_ydjy_sheet[2] = int(ws_ydjy.cell(row = ying_shou_row, column = bi_shu_column).value[7:12].replace(",", ""))
        hui_zong_ydjy_sheet[3] = ws_ydjy.cell(row = ying_shou_row, column = jin_e_column).value[7:]
        hui_zong_ydjy_sheet[4] = int(ws_ydjy.cell(row = ying_fu_row, column = bi_shu_column).value[7:13].replace(",", ""))
        hui_zong_ydjy_sheet[5] = ws_ydjy.cell(row = ying_fu_row, column = jin_e_column).value[7:]
        
        hui_zong_ydjy.append(hui_zong_ydjy_sheet)

#异地平衡
def yi_di_ping_heng(wb_ydph_name, ws_ydph_names, hui_zong_ydph):
    wb_ydph = load_workbook(wb_ydph_name)

    for i, j, p in zip(ws_ydph_names, zb_source_files.work_week.get_week_date(), zb_source_files.work_week.get_week_date_three_day_ago()):
        ws_ydph = wb_ydph[i]
        jin_e_row = 3
        jin_e_column = 2 
        hui_zong_ydph_sheet = [0] * 6
        hui_zong_ydph_sheet[0] = j
        hui_zong_ydph_sheet[1] = p
                
        for q in range(0, 4):
            hui_zong_ydph_sheet[q + 2] = ws_ydph.cell(row = jin_e_row, column = jin_e_column + q).value

        hui_zong_ydph.append(hui_zong_ydph_sheet)

#发卡网点异常流水
def fa_ka(wb_fk_name, ws_fk_names, hui_zong_fk):
    wb_fk = load_workbook(wb_fk_name)

    for i in ws_fk_names:
        l_tag = True
        jin_e_column = 6
        chu_li_column = 14
        ws_fk = wb_fk[i]
        hui_zong_fk_sheet = [0] * 8
        j = 8
        while j <= ws_fk.max_row:
            #已日结
            if ws_fk.cell(row = j, column = jin_e_column).value != None and l_tag:
                if ws_fk.cell(row = j, column = chu_li_column).value[0] == "已":
                    hui_zong_fk_sheet[5] += 1
                    hui_zong_fk_sheet[7] += float(ws_fk.cell(row = j, column = jin_e_column).value)
                hui_zong_fk_sheet[1] += 1
                hui_zong_fk_sheet[3] += float(ws_fk.cell(row = j, column = jin_e_column).value)
            elif ws_fk.cell(row = j, column = jin_e_column).value == None and l_tag:
                j += 3
                l_tag = False
            #X流水
            elif not l_tag:
                if ws_fk.cell(row = j, column = jin_e_column).value != None:
                    if ws_fk.cell(row = j, column = chu_li_column).value[0] == "已":
                        hui_zong_fk_sheet[6] += 1
                        hui_zong_fk_sheet[7] += float(ws_fk.cell(row = j, column = jin_e_column).value)
                    hui_zong_fk_sheet[2] += 1
                    hui_zong_fk_sheet[3] += float(ws_fk.cell(row = j, column = jin_e_column).value)
            j += 1

        hui_zong_fk.append(hui_zong_fk_sheet)


#清算异常流水
def qing_suan(wb_qs_name, ws_qs_names, hui_zong_qs):
    wb_qs = load_workbook(wb_qs_name)

    for i in ws_qs_names:
        l_tag = True
        jin_e_column = 10
        chu_li_column = 36
        ws_qs = wb_qs[i]
        hui_zong_qs_sheet = [0] * 6
        j = 4
        while j <= ws_qs.max_row:
            if ws_qs.cell(row = j, column = jin_e_column).value != None and l_tag:
                if ws_qs.cell(row = j, column = chu_li_column).value == "已处理":
                    hui_zong_qs_sheet[3] += 1
                    hui_zong_qs_sheet[5] += float(ws_qs.cell(row = j, column = jin_e_column).value)
                hui_zong_qs_sheet[0] += 1
                hui_zong_qs_sheet[2] += float(ws_qs.cell(row = j, column = jin_e_column).value)
            elif ws_qs.cell(row = j, column = jin_e_column).value == None and l_tag:
                j += 3
                jin_e_column += 1
                chu_li_column += 3
                l_tag = False
            elif not l_tag:
                if ws_qs.cell(row = j, column = jin_e_column).value != None:
                    if ws_qs.cell(row = j, column = chu_li_column).value == "已处理":
                        hui_zong_qs_sheet[4] += 1
                        hui_zong_qs_sheet[5] += float(ws_qs.cell(row = j, column = jin_e_column).value)
                    hui_zong_qs_sheet[1] += 1
                    hui_zong_qs_sheet[2] += float(ws_qs.cell(row = j, column = jin_e_column).value)
            j += 1

        hui_zong_qs.append(hui_zong_qs_sheet)

if zb_source_files.work_week.diferent_month:
    qing_suan(zb_source_files.last_week_qs, zb_source_files_sheets[:zb_bian_jie_zhi], hui_zong_qs)
    qing_suan(zb_source_files.current_week_qs, zb_source_files_sheets[zb_bian_jie_zhi:], hui_zong_qs)
    zhang_wu_ping_heng(zb_source_files.last_week_ph, zb_source_files_sheets[:zb_bian_jie_zhi], hui_zong_ph)
    zhang_wu_ping_heng(zb_source_files.current_week_ph, zb_source_files_sheets[zb_bian_jie_zhi:], hui_zong_ph)
    yi_di_jiao_yi(zb_source_files.last_week_ydjy, zb_source_files_sheets[:zb_bian_jie_zhi], hui_zong_ydjy)
    yi_di_jiao_yi(zb_source_files.current_week_ydjy, zb_source_files_sheets[zb_bian_jie_zhi:], hui_zong_ydjy)
    yi_di_ping_heng(zb_source_files.last_week_ydph, zb_source_files_sheets[:zb_bian_jie_zhi], hui_zong_ydph)
    yi_di_ping_heng(zb_source_files.current_week_ydph, zb_source_files_sheets[zb_bian_jie_zhi:], hui_zong_ydph)
    fa_ka(zb_source_files.last_week_fk, zb_source_files_sheets[:zb_bian_jie_zhi], hui_zong_fk)
    fa_ka(zb_source_files.current_week_fk, zb_source_files_sheets[zb_bian_jie_zhi:], hui_zong_fk)
else:
    qing_suan(zb_source_files.current_week_qs, zb_source_files_sheets[:zb_bian_jie_zhi], hui_zong_qs)
    zhang_wu_ping_heng(zb_source_files.current_week_ph, zb_source_files_sheets[:zb_bian_jie_zhi], hui_zong_ph)
    yi_di_jiao_yi(zb_source_files.current_week_ydjy, zb_source_files_sheets[:zb_bian_jie_zhi], hui_zong_ydjy)
    yi_di_ping_heng(zb_source_files.current_week_ydph, zb_source_files_sheets[:zb_bian_jie_zhi], hui_zong_ydph)
    fa_ka(zb_source_files.current_week_fk, zb_source_files_sheets[:zb_bian_jie_zhi], hui_zong_fk)


