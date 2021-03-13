from openpyxl import Workbook
from openpyxl import load_workbook
from zhou_bao_files import ZhouBaoSource
#from print_date_of_week import DateOfWeek

zb_source_files = ZhouBaoSource()
current_month_dir = zb_source_files.current_month_dir
zb_source_files_sheets = zb_source_files.work_week.sheet_name_date()
zb_bian_jie_zhi = zb_source_files.work_week.bian_jie_zhi

if zb_source_files.work_week.diferent_month:
    qing_suan(load_workbook(zb_source_files.last_week_qs), zb_source_files_sheets[:zb_bian_jie_zhi])
    qing_suan(load_workbook(zb_source_files.current_week_qs), zb_source_files_sheets[zb_bian_jie_zhi:])
else:
    qing_suan(load_workbook(zb_source_files.current_week_qs, zb_source_files_sheets[:zb_bian_jie_zhi]))

def qing_suan(wb_qs_name, ws_qs_names):
    hui_zong_qs = []
    l_ben_di = []
    l_yi_di = []
    l_tag = True
    jin_e_column = 10
    wb_qs = load_workbook(wb_qs_name)

    for i in ws_qs_names:
        ws_qs = wb_qs[i]

        for j in range(4, ws_qs.max_row + 1):
            if ws_qs.cell(row = i, column = jin_e_column).value != None and l_tag:
                if red:
                    本地未处理笔数 += 1
                    本地未处理金额 += admount
                本地总笔数 += 1
                本地总金额 += admount
            elif l_tag and ws_week.cell(row = i, column = jin_e_column).value != None:
                i += 3
                l_tag = False
                #continue
            elif not l_tag:
                l_yi_di.append(ws_week.cell(row = i, column = jin_e_column).value)

            print(l_ben_di)
            print(l_yi_di)

    return 
